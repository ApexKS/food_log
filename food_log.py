from tkinter import *
from tkinter import ttk
import datetime
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import os.path

root = Tk()
root.geometry('400x200')

mainframe = ttk.Frame(root)
mainframe.pack()

now = datetime.datetime.now()
time = Label(root, text=now.strftime("%d/%m/%Y, %H:%M:%S"))
time.pack()

brf_l = Label(root, text='Breakfast:')
brf_l.pack()
brf_entry = Entry(root)
brf_entry.pack()

lun_l = Label(root, text='Lunch:')
lun_l.pack()
lun_entry = Entry(root)
lun_entry.pack()

din_l = Label(root, text='Dinner:')
din_l.pack()
din_entry = Entry(root)
din_entry.pack()

def Save():
    brf = brf_entry.get()
    lun = lun_entry.get()
    din = din_entry.get()

    new_row = [[now], ["Breakfast", " ", "Lunch", " ", "Dinner"], [brf," ", lun, " ", din]]

    if os.path.isfile("stutee_food_log.xlsx") == False:
        wb = Workbook()
    else:
        wb = load_workbook("stutee_food_log.xlsx")
    ws = wb.active
    for row in new_row:
        ws.append(row)

    col_to_merge = 1
    
    data_to_merge = now

    merge_start = None

    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_to_merge).value

        if cell_value == data_to_merge:
            if merge_start is None:
                merge_start = row
            else:
                if merge_start is not None:
                    ws.merge_cells(start_row=merge_start, end_row=row - 1, start_column=col_to_merge, end_column= col_to_merge)
                    merge_start = None
        
        if merge_start is not None:
            ws.merge_cells(start_row = merge_start, end_row= ws.max_row, start_column=col_to_merge, end_column=col_to_merge)


    wb.save(r'C:\Users\DELL\OneDrive\Desktop\food_log.xlsx')
    messagebox.showinfo("Process complete")

save_but = Button(root, text="Save", command=Save) 
save_but.pack()

root.mainloop()